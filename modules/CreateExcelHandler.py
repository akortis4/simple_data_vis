#module import
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.legend import Legend, LegendEntry
from openpyxl.worksheet.datavalidation import DataValidation

#main class to create excel, format it, write data, and create charts
class CreateExcel():
    def __init__(self, data_frame:pd.DataFrame, stat_data:dict, export:str):
        self.data_frame = data_frame
        self.stat_data = stat_data
        self.export = export
        self.wb = Workbook()
        self.sheet_names = ('Drop Down', 'Stat Data', 'Raw Data', 'Scatter Data')
        self.stat_keys = list(self.stat_data['dep_var'][list(self.stat_data['dep_var'].keys())[0]].keys())
        self.execute()

    def execute(self):
        self.create_tabs()
        self.set_widths()
        self.align_cells()
        self.fill_cells()
        self.bold_font()
        self.outline_cells()
        self.write_dep_values()
        self.write_drop_values()
        self.create_drop_down()
        self.add_vlookups()
        self.write_ind_stats()
        self.write_raw_data()
        self.write_scatter_dep_data()
        self.write_ind_equations()
        self.create_scatter()
        self.hide_sheets()
        self.save_workbook()

    def create_tabs(self):
        ws = self.wb.active
        ws.title = 'Summary Data'
        for sheet_name in self.sheet_names:
            self.wb.create_sheet(sheet_name)

    def set_widths(self):
        ws = self.wb['Summary Data']
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 3
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 3

    def align_cells(self):
        ws = self.wb['Summary Data']
        iterate_len = len(self.stat_keys) + 1
        for i in range(iterate_len):
            ws.cell(row=i+1, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=i+1, column=5).alignment = Alignment(horizontal='center')

    def fill_cells(self):
        ws = self.wb['Summary Data']
        ws['A1'].fill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
        ws['B1'].fill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
        ws['D1'].fill = PatternFill(start_color='DA9694', end_color='DA9694', fill_type='solid')
        ws['E1'].fill = PatternFill(start_color='DA9694', end_color='DA9694', fill_type='solid')
        iter_len = len(self.stat_keys)
        cell_cols = ('A', 'B', 'D', 'E')
        for i in range(iter_len):
            for col in cell_cols:
                cell_value = col + str(i+2)
                if col in ('A', 'B'):
                    ws[cell_value].fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
                else:
                    ws[cell_value].fill = PatternFill(start_color='E6B8B7', end_color='E6B8B7', fill_type='solid')

    def bold_font(self):
        ws = self.wb['Summary Data']
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        ws['D1'].font = Font(bold=True)
        ws['E1'].font = Font(bold=True)

    def outline_cells(self):
        ws = self.wb['Summary Data']
        ws['A1'].border = Border(left=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thin'))
        ws['B1'].border = Border(right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'), left=Side(style='thin'))
        ws['D1'].border = Border(left=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thin'))
        ws['E1'].border = Border(right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'), left=Side(style='thin'))
        iter_len = len(self.stat_keys)
        cell_cols = ('A', 'B', 'D', 'E')
        for i in range(iter_len):
            for col in cell_cols:
                cell_value = col + str(i+2)
                if col in ('A', 'D'):
                    if i == (iter_len - 1):
                        ws[cell_value].border = Border(left=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thin'))
                    else:
                        ws[cell_value].border = Border(left=Side(style='thick'), bottom=Side(style='thin'), right=Side(style='thin'))
                else:
                    if i == (iter_len - 1):
                        ws[cell_value].border = Border(right=Side(style='thick'), bottom=Side(style='thick'))
                    else:
                        ws[cell_value].border = Border(right=Side(style='thick'), bottom=Side(style='thin'))

    def write_dep_values(self):
        ws = self.wb['Summary Data']
        dep_var = list(self.stat_data['dep_var'].keys())[0]
        ws.cell(row=1, column=1).value = 'Dependent Variable:'
        ws.cell(row=1, column=2).value = dep_var
        for x, (key, value) in enumerate(self.stat_data['dep_var'][dep_var].items()):
            ws.cell(row=2+x, column=1).value = key
            ws.cell(row=2+x, column=2).value = value

    def write_drop_values(self):
        ws = self.wb['Drop Down']
        ind_vars = list(self.stat_data['ind_var'].keys())
        ind_vars.sort()
        for x, i in enumerate(ind_vars):
            ws.cell(row=x+1, column=1).value = i

    def create_drop_down(self):
        ws = self.wb['Summary Data']
        ws.cell(row=1, column=4).value = 'Independent Variable:'
        for x, key in enumerate(self.stat_keys):
            ws.cell(row=2+x, column=4).value = key
        data_len = int(self.data_frame.shape[0])
        data_validation = DataValidation(type='list', formula1=f"='Drop Down'!A1:A{data_len}")
        ws.add_data_validation(data_validation)
        data_validation.add(ws['E1'])

    def add_vlookups(self):
        ws = self.wb['Summary Data']
        for x, key in enumerate(self.stat_keys):
            ws.cell(row=2+x, column=5).value = f'=iferror(vlookup(E1,\'Stat Data\'!A:I,{2+x},False),"")'
        
    def write_ind_stats(self):
        ws = self.wb['Stat Data']
        for x, (key, value) in enumerate(self.stat_data['ind_var'].items()):
            ws.cell(row=1+x, column=1).value = key
            for y, stat_key in enumerate(self.stat_keys):
                ws.cell(row=1+x, column=2+y).value = value[stat_key]

    def write_raw_data(self):
        ws = self.wb['Raw Data']
        for x, key in enumerate(self.stat_data['ind_var'].keys()):
            values = list(self.data_frame[key].values)
            ws.cell(row=1, column=1+x).value = key
            for y, value in enumerate(values):
                ws.cell(row=2+y, column=1+x).value = value

    def write_scatter_dep_data(self):
        ws = self.wb['Scatter Data']
        dep_var = list(self.stat_data['dep_var'].keys())[0]
        dep_values = list(self.data_frame[dep_var].values)
        ws['A1'] = dep_var
        for x, value in enumerate(dep_values):
            ws.cell(row=2+x, column=1).value = value

    def write_ind_equations(self):
        ws = self.wb['Scatter Data']
        ws['B1'] = '=\'Summary Data\'!E1'
        for i in range(int(self.data_frame.shape[0])):
            cell_value = f'=iferror(hlookup(B1, \'Raw Data\'!A:AJ, {i+2}, False), "")'
            ws.cell(row=i+2, column=2).value = cell_value

    def create_scatter(self):
        ws = self.wb['Summary Data']
        scat_ws = self.wb['Scatter Data']
        chart = ScatterChart()
        chart.title = 'Scatter Plot'
        xvalues = Reference(scat_ws, min_row=2, max_row=int(self.data_frame.shape[0]), min_col=2)
        yvalues = Reference(scat_ws, min_row=2, max_row=int(self.data_frame.shape[0]), min_col=1)
        series_data = Series(yvalues, xvalues, title_from_data=False)
        chart.series.append(series_data)
        chart.legend = None
        s1 = chart.series[0]
        s1.marker.symbol = "triangle"
        s1.graphicalProperties.line.noFill = True
        ws.add_chart(chart, 'G1')

    def hide_sheets(self):
        for sheet in self.sheet_names:
            self.wb[sheet].sheet_state = 'hidden'

    def save_workbook(self):
        self.wb.save(self.export + '/1_1_summary_data.xlsx')